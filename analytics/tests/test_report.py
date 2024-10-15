import os
from django.test import TestCase
from django.utils import timezone
from analytics.models import Category, Customer, Inventory, Order, OrderItem, Product

from analytics.reports import export_monthly_sales_report


class MonthlySalesReportTest(TestCase):
     
    def setUp(self):
        self.category = Category.objects.create(name='Electronics')
        self.product = Product.objects.create(
            name='Test Product',
            description='A high-performance laptop.',
            SKU='LAP123',
            price=10.0,
            category=self.category
        )

        self.inventory, created = Inventory.objects.get_or_create(
            product=self.product,
            defaults={'quantity': 100}
        )

        self.customer = Customer.objects.create(
            name='John Doe',
            email='john.doe@example.com',
            country='US'
        )
        self.order = Order.objects.create(
            customer=self.customer,
            order_date=timezone.now().date(),
            status='C',
            total_amount=10.0
        )

        # Create an order and order items
        OrderItem.objects.create(order=self.order, product=self.product, quantity=5, price_at_time_of_order=self.product.price)


    def test_export_monthly_sales_report(self):
        # Call the function to export the report
        file_path = export_monthly_sales_report(10, 2024)
        
        # Check if the file was created
        self.assertTrue(os.path.exists(file_path))

        # Verify the contents of the file (optional)
        from openpyxl import load_workbook
        
        wb = load_workbook(file_path)
        ws = wb.active

        # Check header
        self.assertEqual(ws['A1'].value, "Product")
        self.assertEqual(ws['B1'].value, "Quantity")
        self.assertEqual(ws['C1'].value, "Price")
        self.assertEqual(ws['D1'].value, "Total Sales")
        self.assertEqual(ws['E1'].value, "Remaining Quantity")

        # Check data row
        self.assertEqual(ws['A2'].value, self.product.name)
        self.assertEqual(ws['B2'].value, 5)
        self.assertEqual(ws['C2'].value, self.product.price)
        self.assertEqual(ws['D2'].value, 50.0)  
        self.assertEqual(ws['E2'].value, 95) 

        # Clean up the created file
        os.remove(file_path)

    def tearDown(self):
        # Clean up test data if needed
        Inventory.objects.all().delete()
        OrderItem.objects.all().delete()
        Order.objects.all().delete()